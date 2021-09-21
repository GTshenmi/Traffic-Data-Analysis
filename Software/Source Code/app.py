import numpy as np
import openpyxl as xl
from openpyxl.styles import *
import random
import datetime

ON = 1
OFF = 0
DEBUG = ON


class Performance():
    def __init__(self):
        self.MAE = 0.0
        self.MAPE = 0.0
        self.RMSE = 0.0
        self._MAE = 0.0
        self._MAPE = 0.0
        self._RMSE = 0.0


def Main():
    workbook = xl.load_workbook("北京快速路2020-10-27 16-50-30.668500.xlsx")
    filename = DataPreProcessSheet1(workbook)
    workbook = xl.load_workbook(filename)
    PerformanceEvaluation(workbook)


def DataRemoval(workbook):
    for sheet in workbook:

        max_rows = sheet.max_row
        max_cols = sheet.max_column
        #抽取500行
        rows = random.sample(sheet[2:max_rows + 1], 500)

        random_sheet = workbook.create_sheet(f'random_sheet_{str(sheet.title)}')
        #将500行数据写入新工作簿
        for i in range(0, max_cols):
            value = sheet.cell(1, i + 1).value
            random_sheet.cell(1, i + 1, value)

        for i in range(0, 500):
            for j in range(0, max_cols):
                value = rows[i][j].value
                random_sheet.cell(i + 2, j + 1, value)

    time = str(datetime.datetime.now()).replace(":", "-")
    #保存
    filename = f'北京快速路{time}.xlsx'
    workbook.save(filename)

    workbook.close()
    return filename


def DataCompletion(Order, last_row, now_row):
    # 计算Y矩阵
    last_value = last_row.value
    now_value = now_row.value
    Y = []
    if Order == 1:
        Y.append((last_value + now_value) / 2)
    else:
        Y.append(last_value)
        for i in range(Order - 2):
            Y.append(0)
        Y.append(now_value)

    if DEBUG == ON:
        print("Y矩阵:")
        print(Y)

    # 计算系数矩阵
    CoefficientMatrix = CreateCoefficientMatrix(Order)

    if DEBUG == ON:
        print("系数矩阵:")
        print(CoefficientMatrix)

    Matrix = np.array(CoefficientMatrix)

    # 求系数矩阵的逆矩阵
    InverseMatrix = np.linalg.inv(Matrix)

    if DEBUG == ON:
        print("逆矩阵:")
        print(InverseMatrix)

    # 解方程组：  CoefficientMatrix .* ResultMatrix = Y
    ResultMatrix = np.dot(InverseMatrix, Y)

    # 结果取整
    ResultMatrixInt = []
    for result in ResultMatrix:
        result = int(result + 0.5)
        ResultMatrixInt.append(result)

    if DEBUG == ON:
        print("结果矩阵:")
        print(ResultMatrixInt)

    return ResultMatrixInt


def DataPreProcessSheet1(workbook):
    names = workbook.get_sheet_names()

    sheet = workbook.get_sheet_by_name(names[3])

    lastcell = sheet['C2']

    ReductionSheet = workbook.create_sheet(f'Reduction_{sheet.title}')

    row_count = 0
    row_count_new = 0
    #遍历表格C列
    for cell in sheet['C']:
        if cell.value is not None and str(cell.value) != 'TIME':
            #计算相邻表格时间差，判断有无数据缺失
            dt = GetTimeDif(lastcell.value, cell.value)
            if dt != 2 and dt != -1:
                now_row = sheet[cell.row]
                last_row = sheet[lastcell.row]
                order = int(dt / 2 - 1)
                # 有数据缺失，生成补全数据，写入新工作簿
                rows = GenerateRowData(order=order, last_row=last_row, now_row=now_row, count=row_count)

                for row in rows:
                    row_count_new += 1
                    for i in range(0, 6):
                        ReductionSheet.cell(row_count_new, i + 1).font = Font(color='FF0000')
                        ReductionSheet.cell(row_count_new, i + 1, row[i])

                row_count_new += 1
                row_count += 1
                for i in range(6):
                    ReductionSheet.cell(row_count_new, i + 1, sheet.cell(row_count, i + 1).value)

            else:
                row_count += 1
                row_count_new += 1
                for i in range(6):
                    ReductionSheet.cell(row_count_new, i + 1, sheet.cell(row_count, i + 1).value)

            lastcell = cell
        else:
            row_count += 1
            row_count_new += 1
            for i in range(6):
                ReductionSheet.cell(row_count_new, i + 1, sheet.cell(row_count, i + 1).value)

    for cell in ReductionSheet['A']:
        cell.value = cell.row - 1
    ReductionSheet['A1'] = 'ID'

    time = str(datetime.datetime.now()).replace(":", "-")

    filename = f'北京快速路{time}.xlsx'
    workbook.save(filename)
    workbook.close()
    return filename


def PerformanceEvaluation(workbook):
    names = workbook.get_sheet_names()
    raw_sheet = workbook.get_sheet_by_name(names[0])
    comple_sheet = workbook.get_sheet_by_name(names[6])

    raw_cells = raw_sheet['D2':'F715']
    comple_cells = comple_sheet['D2':'F715']

    L = [Performance(), Performance(), Performance()]

    for j in range(3):
        raw_index = 0
        comple_index = 0
        #遍历新表
        for i in range(len(comple_cells)):
            if comple_index >= len(comple_cells):
                break
            raw_cell = raw_cells[raw_index][j]
            comple_cell = comple_cells[comple_index][j]
            #和原表相应单元格作差，得出误差
            if CellIsCorresponding(raw_sheet[raw_cell.row][2].value, comple_sheet[comple_cell.row][2].value):
                pass
            else:
                comple_index += 1
                continue

            raw_index += 1
            comple_index += 1
            Xi = int(raw_cell.value)
            _Xi = int(comple_cell.value)

            L[j]._MAE += abs((Xi - _Xi))
            if Xi != 0:
                L[j]._MAPE += abs((Xi - _Xi) / Xi)
            else:
                L[j]._MAPE += 0

    for j in range(3):
        L[j].MAE = L[j]._MAE / float(len(raw_cells))
        L[j].MAPE = L[j]._MAPE / float(len(raw_cells))
        L[j].RMSE = np.sqrt(L[j].MAE)

    #将性能评估数据写入新表
    comple_sheet.cell(716 + 0, 2, " ")
    comple_sheet.cell(716 + 0, 1, "MAE")
    comple_sheet.cell(716 + 1, 1, "MAPE")
    comple_sheet.cell(716 + 2, 1, "RMSE")

    for i in range(0, 3):
        comple_sheet.cell(716, 4+ i, L[i].MAE)
        comple_sheet.cell(717, 4+ i, L[i].MAPE)
        comple_sheet.cell(718, 4+ i, L[i].RMSE)

    for i in range(3):
        print(f'L{i}.MAE:{L[i].MAE}')
        print(f'L{i}.MAPE:{L[i].MAPE}')
        print(f'L{i}.RMSE:{L[i].RMSE}')

    time = str(datetime.datetime.now()).replace(":", "-")

    filename = f'北京快速路{time}.xlsx'
    workbook.save(filename)
    workbook.close()
    return filename


def GetTimeDif(last, now):
    if last != now:

        last_hour, last_min, last_sec = str(last).split(':')
        now_hour, now_min, now_sec = str(now).split(':')
        if DEBUG == ON:
            print("上一时刻:")
            print(last)

        if DEBUG == ON:
            print("当前时刻:")
            print(now)
        return int((int(now_hour) - int(last_hour)) * 60 + (int(now_min) - int(last_min)))
    else:
        return -1


def GetCellTime(last_time, err):
    last_hour, last_min, last_sec = str(last_time).split(':')

    timebase = int(last_hour) * 3600 + int(last_min) * 60 + int(last_sec)

    time = int(timebase + 2 * err * 60)

    hour = int(time / 3600)
    min = int(time % 3600 / 60)
    sec = int(time % 60)

    result = datetime.time(hour, min, sec)

    return result


def CellIsCorresponding(time1, time2):
    time1_hour, time1_min, time1_sec = str(time1).split(':')
    time2_hour, time2_min, time2_sec = str(time2).split(':')
    if time1_hour == time2_hour and time1_min == time2_min:
        return True
    else:
        return False


def GenerateRowData(order, last_row, now_row, count):
    rows = []
    ResultMatrixs = []

    for i in range(3, 6):
        ResultMatrix = DataCompletion(Order=order, last_row=last_row[i], now_row=now_row[i])
        ResultMatrixs.append(ResultMatrix)
    ResultMatrixs = np.reshape(ResultMatrixs, (3, order))

    for i in range(order):
        for j in range(6):
            if j == 0:
                rows.append(count + i)
            elif j == 1:
                rows.append(now_row[1].value)
            elif j == 2:
                rows.append(GetCellTime(last_row[2].value, i + 1))
            else:
                rows.append(ResultMatrixs[j - 3][i])
    rows = np.reshape(rows, (order, 6))

    return rows


def CreateCoefficientMatrix(Order):
#    @example:Order = 7:
#    [[ 2 -1  0  0  0  0  0]
#     [-1  2 -1  0  0  0  0]
#     [ 0 -1  2 -1  0  0  0]
#     [ 0  0 -1  2 -1  0  0]
#     [ 0  0  0 -1  2 -1  0]
#     [ 0  0  0  0 -1  2 -1]
#     [ 0  0  0  0  0  -1 2]]
    Matrix = []

    if Order == 1:
        Matrix.append(1)
    else:
        for i in range(Order):
            for j in range(Order):
                if i == 0:
                    if j == 0:
                        Matrix.append(2)
                    elif j == 1:
                        Matrix.append(-1)
                    else:
                        Matrix.append(0)
                elif i == Order - 1:
                    if j == Order - 1:
                        Matrix.append(2)
                    elif j == Order - 2:
                        Matrix.append(-1)
                    else:
                        Matrix.append(0)
                else:
                    if j == i - 1:
                        Matrix.append(-1)
                    elif j == i:
                        Matrix.append(2)
                    elif j == i + 1:
                        Matrix.append(-1)
                    else:
                        Matrix.append(0)

    if DEBUG == ON:
        print("系数矩阵:Matrix")
        print(Matrix)

    ReshapeMatrix = np.reshape(Matrix, (Order, Order))

    if DEBUG == ON:
        print("系数矩阵:ReshapeMatrix")
        print(ReshapeMatrix)

    return ReshapeMatrix


if __name__ == '__main__':
    try:
        Main()
    except Exception as e:
        if DEBUG == ON:
            print("Error:")
            print(e)
